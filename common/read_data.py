"""
@author:xue
@time:2020/8/6
@desc:读取数据
"""
import csv
import os

import yaml
from openpyxl import load_workbook

from config.conf import BASE_DIR


def join_path(path):
    return os.path.join(BASE_DIR, path).replace('/', '\\')


def read_excel(excel_path, sheet_name):
    wb = load_workbook(excel_path)
    # print(wb.sheetnames)
    sheet = wb[sheet_name]  # wb.get_sheet_by_name(sheet_name)
    # print(type(sheet.rows))

    max_col = sheet.max_column
    data_list = []
    #
    for line in list(sheet.rows)[1:]:  # 跳过第一行
        temp_list = []
        for col in range(1, max_col):  # 跳过第一列
            temp_list.append(line[col].value)
        data_list.append(temp_list)
    return data_list


def read_yaml(yaml_path):
    arr = []
    with open(yaml_path, 'r') as f:
        data = yaml.safe_load(f).values()
        for row in data:
            arr.append(tuple(row.values()))

    print(','.join(list(data)[0].keys()), arr)
    # return ','.join(list(data)[0].keys()), arr
    return arr


def read_csv(csv_path):
    arr = []
    with open(csv_path, 'r') as f:
        data = csv.reader(f)
        for row in data:
            arr.append(tuple(row))
        return ','.join(arr[0]), arr[1:]


if __name__ == '__main__':
    print(join_path('data/login.yml'))
    # print(read_csv(join_path('data/login.csv')))
    print(read_yaml(join_path('data/login.yml')))

    # @pytest.mark.parametrize("name,pwd", [("yy1", "123"), ("yy2", "123")])
